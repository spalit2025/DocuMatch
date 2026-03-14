"""
Export System for DocuMatch Architect.

Generates downloadable reports in PDF and Excel formats.

Supported Formats:
  - Excel (.xlsx): Tabular results with summary sheet + details sheet
  - Text (.txt):   Human-readable validation reports (existing)
"""

import io
import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .database import Database, Result
from .models import ThreeWayMatchResult

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="F8FAFC", bold=True, size=11)
PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

STATUS_FILLS = {"PASS": PASS_FILL, "FAIL": FAIL_FILL, "REVIEW": REVIEW_FILL}


def export_results_excel(
    database: Database,
    vendor_name: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 500,
) -> bytes:
    """
    Export validation results to Excel format.

    Creates a workbook with:
    - Summary sheet: KPI metrics
    - Results sheet: Individual validation results

    Args:
        database: Database instance for querying results
        vendor_name: Optional vendor filter
        status: Optional status filter (PASS/FAIL/REVIEW)
        limit: Max results to export

    Returns:
        Excel file bytes (write to disk or return as HTTP response)
    """
    wb = Workbook()

    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"

    stats = database.get_stats()

    # Title
    ws_summary["A1"] = "DocuMatch Validation Report"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary["A2"].font = Font(color="666666", size=10)

    # KPI section
    kpis = [
        ("Total Processed", stats["total_results"]),
        ("Pass Rate", f"{stats['pass_rate']:.0%}"),
        ("Passed", stats["pass_count"]),
        ("Failed", stats["fail_count"]),
        ("Review", stats["review_count"]),
        ("Total Jobs", stats["total_jobs"]),
        ("Completed Jobs", stats["completed_jobs"]),
        ("Failed Jobs", stats["failed_jobs"]),
    ]

    for i, (label, value) in enumerate(kpis, start=4):
        ws_summary[f"A{i}"] = label
        ws_summary[f"A{i}"].font = Font(bold=True)
        ws_summary[f"B{i}"] = value

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    # ===== RESULTS SHEET =====
    ws_results = wb.create_sheet("Results")

    headers = [
        "Invoice #", "Vendor", "Status", "Confidence",
        "Matches Passed", "Total Matches", "File", "Date",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Query results
    results = database.get_results(
        vendor_name=vendor_name, status=status, limit=limit,
    )

    for row_idx, r in enumerate(results, start=2):
        ws_results.cell(row=row_idx, column=1, value=r.invoice_number or "N/A")
        ws_results.cell(row=row_idx, column=2, value=r.vendor_name or "N/A")

        status_cell = ws_results.cell(row=row_idx, column=3, value=r.status or "N/A")
        if r.status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[r.status]

        ws_results.cell(
            row=row_idx, column=4,
            value=f"{r.confidence:.0%}" if r.confidence else "N/A",
        )
        ws_results.cell(row=row_idx, column=5, value=r.matches_passed)
        ws_results.cell(row=row_idx, column=6, value=r.total_matches)
        ws_results.cell(row=row_idx, column=7, value=r.invoice_file or "N/A")
        ws_results.cell(
            row=row_idx, column=8,
            value=r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "N/A",
        )

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws_results.column_dimensions[
            ws_results.cell(row=1, column=col).column_letter
        ].width = 18

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
